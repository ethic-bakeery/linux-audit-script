import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

# Define the directory containing JSON files
JSON_DIR = "."

# Define the output Excel file
OUTPUT_EXCEL = "final_audit_report.xlsx"

# Function to load JSON data from a file
def load_json(file_path):
    with open(file_path, "r") as file:
        try:
            # Try to load as a single JSON object
            return json.load(file)
        except json.JSONDecodeError:
            # If it fails, try to load as multiple JSON objects
            file.seek(0)  # Reset file pointer
            data = []
            for line in file:
                line = line.strip()
                if line:
                    data.append(json.loads(line))
            return data

# Function to create an Excel report
def create_excel_report(json_files):
    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Audit Report"

    # Define styles for the header and data rows
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write the title of the report
    worksheet.merge_cells("A1:C1")
    title_cell = worksheet.cell(row=1, column=1, value="Final Audit Report")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Start writing data from row 3
    current_row = 3

    # Process each JSON file
    for json_file in json_files:
        file_path = os.path.join(JSON_DIR, json_file)
        data = load_json(file_path)

        # Extract the section title from the file name (without .json)
        section_title = json_file.replace(".json", "").replace("_", " ").title()

        # Write the section header
        worksheet.merge_cells(f"A{current_row}:C{current_row}")
        section_cell = worksheet.cell(row=current_row, column=1, value=f"Audit Results: {section_title}")
        section_cell.font = Font(bold=True, size=14)
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        # Write the table headers
        headers = ["Check", "Status", "Recommendation"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        current_row += 1

        # Extract the audit results
        audit_results = []
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list):
                    audit_results.extend(value)
                else:
                    audit_results.append(value)
        elif isinstance(data, list):
            audit_results.extend(data)

        # Write the audit results to the worksheet
        for result in audit_results:
            if isinstance(result, dict):
                check = result.get("check", "N/A")
                status = result.get("status", "N/A")
                recommendation = result.get("recommendation", "N/A")

                # Write the row
                worksheet.cell(row=current_row, column=1, value=check).alignment = data_alignment
                worksheet.cell(row=current_row, column=2, value=status).alignment = data_alignment
                worksheet.cell(row=current_row, column=3, value=recommendation).alignment = data_alignment

                # Apply borders to the row
                for col_num in range(1, 4):
                    worksheet.cell(row=current_row, column=col_num).border = thin_border

                current_row += 1

        # Add a blank row between sections
        current_row += 1

    # Adjust column widths (Fix for MergedCell error)
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            if isinstance(cell, MergedCell):  # Skip merged cells
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    workbook.save(OUTPUT_EXCEL)
    print(f"Excel report generated: {OUTPUT_EXCEL}")

# List of JSON files to include in the report
json_files = [
    "apparmor_audit_report.json",
    "chrony_audit.json",
    "cli_warning_banners_audit.json",
    "data_retention_audit.json",
    "filesystem_audit_report.json",
    "fips_audit.json",
    "gdm_security_audit.json",
    "host_based_firewall_audit.json",
    "iptables_audit.json",
    "local_user_group_audit.json",
    "log_file_access_audit.json",
    "logging_audit.json",
    "network_devices_audit.json",
    "network_kernel_modules_audit.json",
    "network_kernel_parameters_audit.json",
    "nftables_audit.json",
    "pam_pkcs11_audit.json",
    "pam_pwquality_audit.json",
    "partition_audit_report.json",
    "password_policy_audit.json",
    "privilege_escalation_audit.json",
    "rsyslog_audit.json",
    "secure_boot_audit_report.json",
    "service_clients_audit.json",
    "software_patch_audit_report.json",
    "special_services_audit.json",
    "ssh_server_audit.json",
    "timesyncd_audit.json",
    "time_sync_audit.json",
    "user_accounts_audit.json",
]

# Generate the Excel report
create_excel_report(json_files)
